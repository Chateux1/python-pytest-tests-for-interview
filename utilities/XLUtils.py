import openpyxl


def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def getColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def readData(file, sheetname, rownum, column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=column).value


def writeData(file, sheetname, rownum, column, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=column).value = data
    workbook.save(file)
